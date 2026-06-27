"""Generate realistic sample RFQ files to demo multimodal ingestion.
Outputs: sample_rfq.pdf (customer purchase inquiry) + sample_rfq.xlsx (messy table)."""
import os
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))

# ---------------- PDF: a gold-mine customer RFQ email/letter ----------------
def make_pdf():
    doc = SimpleDocTemplate(os.path.join(HERE, "sample_rfq.pdf"), pagesize=A4,
                            topMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm)
    ss = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=ss["Title"], fontSize=13, spaceAfter=2)
    n = ss["Normal"]; n.fontSize = 9.5; n.leading = 13
    small = ParagraphStyle("s", parent=n, fontSize=8.5, textColor=colors.grey)
    el = []
    el += [Paragraph("PT BUMI SUKSESINDO", h),
           Paragraph("Tujuh Bukit Gold Mine — Banyuwangi, Jawa Timur", small),
           Spacer(1, 10),
           Paragraph("<b>REQUEST FOR QUOTATION</b>", n),
           Paragraph("RFQ No: BSI/PROC/2026/0456 &nbsp;&nbsp; Date: 25 June 2026", n),
           Paragraph("To: Kemindo Group (marketing@kemindogroup.com)", n),
           Spacer(1, 8),
           Paragraph("Dear Sir/Madam,", n),
           Paragraph("Please provide your best quotation for the following process "
                     "chemicals for our CIL plant on a <b>monthly contract</b> basis:", n),
           Spacer(1, 8)]
    data = [["No", "Material Description", "Qty", "Unit", "Remarks"],
            ["1", "Hydrated Lime Ca(OH)2", "200", "MT", "min 90% purity, pH control"],
            ["2", "Activated Carbon (coconut shell)", "40", "MT", "iodine no. >= 1000"],
            ["3", "Sodium Metabisulphite (SMBS)", "25", "MT", "cyanide detox"],
            ["4", "Caustic Soda Flake", "30", "MT", "98% min"],
            ["5", "Copper Sulphate", "8", "MT", "flotation activator"]]
    t = Table(data, colWidths=[10*mm, 60*mm, 14*mm, 14*mm, 50*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#22303f")),
        ("TEXTCOLOR", (0,0),(-1,0), colors.white),
        ("FONTSIZE", (0,0),(-1,-1), 8.5),
        ("GRID", (0,0),(-1,-1), 0.4, colors.HexColor("#cfd6df")),
        ("ALIGN", (2,0),(3,-1), "CENTER"),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, colors.HexColor("#f4f6f8")]),
        ("TOPPADDING", (0,0),(-1,-1), 4), ("BOTTOMPADDING", (0,0),(-1,-1), 4)]))
    el += [t, Spacer(1, 10),
           Paragraph("Delivery term: <b>CIF Banyuwangi port</b>", n),
           Paragraph("Payment term: <b>30 days after delivery (NET30)</b>", n),
           Paragraph("Required delivery: within 3 weeks from PO", n),
           Paragraph("Please include MSDS and technical datasheet for each item.", n),
           Spacer(1, 10),
           Paragraph("Best regards,", n),
           Paragraph("Procurement Department<br/>PT Bumi Suksesindo", n)]
    doc.build(el)
    print("wrote sample_rfq.pdf")


# ---------------- XLSX: a messier spreadsheet RFQ ----------------
def make_xlsx():
    wb = Workbook(); ws = wb.active; ws.title = "RFQ"
    ws["A1"] = "PT Indah Kiat Pulp & Paper - RFQ"
    ws["A2"] = "Contact: purchasing@ikpp.example  | Term: NET60 | EXW"
    rows = [["item", "product", "qty", "uom", "note"],
            [1, "OBA optical brightener", 5, "MT", "brightness improvement machine 2"],
            [2, "Retention aid (cationic PAM)", 12, "MT", "drainage drop"],
            [3, "Bentonite retention StarLITE", 8, "MT", ""],
            [4, "ASA sizing", 6, "MT", "internal sizing"]]
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(HERE, "sample_rfq.xlsx"))
    print("wrote sample_rfq.xlsx")


if __name__ == "__main__":
    make_pdf()
    make_xlsx()
